import os
import re
from pathlib import Path
from openpyxl import load_workbook
from dateutil.parser import parse
from config import (
    FILENAME_TO_MASTER_NAME, 
    DIRECTORY_PATHS, 
    IMPORTANT_CASHEET_DATA_COL, 
    LOCATION_START_COL
)

# Global report dictionary to track results
report = {
    'success': [],  # [(location, date, filename)]
    'errors': []    # [(location, date, filename, error_message)]
}

# -----------------------------------------------------------------------------
# 📄 Casheet Helper Functions
# -----------------------------------------------------------------------------

def get_casheet_date(casheet_ws):
    """
    Scans the first 5 rows of the casheet for a cell containing 'Date:'.
    Returns the value of the cell immediately to the right.
    """
    for row in casheet_ws.iter_rows(min_row=1, max_row=5):
        for i in range(len(row) - 1):  # -1 to avoid index out of range at the end
            cell_val = str(row[i].value).strip() if row[i].value else ""
            
            # Case-insensitive check for "Date:"
            if "date:" in cell_val.lower():
                date_value = row[i + 1].value
                if date_value:
                    try:
                        # Use fuzzy=True to handle string variations
                        return parse(str(date_value), fuzzy=True).date()
                    except ValueError:
                        continue
    return None

def get_casheet_data(casheet_ws, keyword_to_find):
    """
    Scans Column C (likely index 2 or 3 depending on merge) for the keyword.
    If found, extracts the specific columns defined in config.
    """
    # Assuming Column C is index 3 (openpyxl is 1-based)
    target_col_letter = "B"  # Change to "C" if needed
    
    # We scan rows 4 to 60 to be safe/fast
    for row_idx in range(4, 60):
        cell_val = casheet_ws[f"{target_col_letter}{row_idx}"].value
        if cell_val and keyword_to_find.lower() in str(cell_val).lower():
            # Keyword found! Extract the important columns from this row
            data = []
            for col_idx in IMPORTANT_CASHEET_DATA_COL:
                val = casheet_ws.cell(row=row_idx, column=col_idx).value
                data.append(val if val is not None else 0)
            return data
            
    return []

# -----------------------------------------------------------------------------
# 📊 Breakdown Master Helper Functions
# -----------------------------------------------------------------------------

def find_master_row_by_date(tender_ws, target_date):
    """
    Scans Column A of the Master Breakdown for the matching date.
    """
    # Start at row 3 (assuming headers are 1-2)
    for row in tender_ws.iter_rows(min_row=3, max_col=1):
        cell_obj = row[0]
        if cell_obj.value:
            try:
                # Convert cell value to date for comparison
                cell_date = parse(str(cell_obj.value), fuzzy=True).date()
                if cell_date == target_date:
                    return cell_obj.row
            except ValueError:
                continue
    return None

def fill_tender_breakdown(tender_ws, data_values, row_idx, start_col_idx):
    """
    Writes the extracted values into the Master Breakdown.
    """
    for i, value in enumerate(data_values):
        # start_col_idx + i gives us the specific tender column (Sales, Flex, etc.)
        tender_ws.cell(row=row_idx, column=start_col_idx + i, value=value)

# -----------------------------------------------------------------------------
# ⚙️ Core Processing Logic
# -----------------------------------------------------------------------------

def process_single_casheet(casheet_path, master_ws):
    """
    1. Identifies the location(s) based on filename.
    2. Extracts data for each location found in the file.
    3. Fills the Master Worksheet object (in memory).
    """
    filename = casheet_path.name
    
    # 1. Clean filename to find the "Key" (e.g., "Crimson Corner 11-13-25" -> "crimson corner")
    # This Regex removes date patterns at the end
    clean_name = re.sub(r'\s*\d{1,2}-\d{1,2}-\d{2,4}.*', '', casheet_path.stem).strip().lower()
    
    # Get the list of mappings for this file
    mappings = FILENAME_TO_MASTER_NAME.get(clean_name)
    
    if not mappings:
        report['errors'].append(("Unknown", None, filename, f"No config mapping found for key: '{clean_name}'"))
        return

    # Load the Casheet File
    try:
        all_wb_casheet = load_workbook(casheet_path, data_only=True)
        # Assuming data is on the first sheet or a sheet matching the day name            
    except Exception as e:
        report['errors'].append(("File Error", None, filename, str(e)))
        return
    
    # loop through all worksheets in the casheet workbook
    for ws_casheet in all_wb_casheet.worksheets:
        # if not ws_casheet.sheet_state == 'visible' or ws_casheet.title.lower() == 'totals':
        #     continue

        # Get Date from Casheet
        report_date = get_casheet_date(ws_casheet)
        if not report_date:
            report['errors'].append(("General", ws_casheet.title, filename, "Could not find valid DATE in file header"))
            return

        # Find the matching Date Row in Master (do this once per file to save time)
        master_row = find_master_row_by_date(master_ws, report_date)
        if not master_row:
            report['errors'].append(("General", report_date, filename, "Date not found in Master Breakdown file"))
            return

        # Iterate through each location mapped to this file
        # Example: Crimson Corner file might have mappings for [{"Crimson Corner": "Crimson Corner"}, {"Gardner": "Thirst"}]
        for mapping_dict in mappings:
            for master_loc_name, casheet_keyword in mapping_dict.items():
                
                # Check if we have a column config for this master location
                start_col = LOCATION_START_COL.get(master_loc_name.lower())
                if not start_col:
                    report['errors'].append((master_loc_name, report_date, filename, "Master Column start index not configured in config.py"))
                    continue

                # Extract Data
                data = get_casheet_data(ws_casheet, casheet_keyword)
                
                if data:
                    # Write Data to Master (In Memory)
                    fill_tender_breakdown(master_ws, data, master_row, start_col)
                    report['success'].append((master_loc_name, report_date, filename))
                else:
                    report['errors'].append((master_loc_name, report_date, filename, f"Keyword '{casheet_keyword}' not found in rows"))

# -----------------------------------------------------------------------------
# 🚀 Main Execution
# -----------------------------------------------------------------------------

def main():
    casheet_dir = Path(DIRECTORY_PATHS["casheets_dir"])
    master_path = Path(DIRECTORY_PATHS["master_path"])

    # Validations
    if not casheet_dir.exists():
        print(f"❌ Error: Casheet directory not found at {casheet_dir}")
        return
    if not master_path.exists():
        print(f"❌ Error: Master Breakdown file not found at {master_path}")
        return

    # Get all Excel files
    files = list(casheet_dir.glob("*.xlsx"))
    if not files:
        print("⚠️ No .xlsx files found in casheet directory.")
        return

    print(f"📂 Found {len(files)} files. Starting process...")

    # Load Master Breakdown ONCE
    try:
        wb_master = load_workbook(master_path)
        ws_master = wb_master.active
    except Exception as e:
        print(f"❌ Critical Error loading Master Breakdown: {e}")
        return

    # Process each file
    for file_path in files:
        # We assume files starting with ~$ are temporary lock files and skip them
        if file_path.name.startswith("~$"):
            continue
            
        process_single_casheet(file_path, ws_master)

    # Save Master Breakdown ONCE at the end
    try:
        wb_master.save(master_path)
        print("💾 Master Breakdown saved successfully!")
    except PermissionError:
        print("\n❌ ERROR: Could not save the Master Breakdown file.")
        print("👉 Please close 'Tender Breakdown.xlsx' and try again.")
        return

    # Generate Final Report
    print("\n" + "="*60)
    print("📋 AUTOFILL SUMMARY REPORT")
    print("="*60)
    
    print(f"\n✅ Successful Updates: {len(report['success'])} / {len(report['success']) + len(report['errors'])}")
    for loc, date, fname in report['success']:
        print(f"   • {loc:<20} | {date} | Source: {fname}")

    print(f"\n⚠️ Errors / Warnings: {len(report['errors'])}")
    for loc, date, fname, err in report['errors']:
        d_str = str(date) if date else "N/A"
        print(f"   • {loc if loc else 'Unknown':<20} | {d_str:<10} | {fname}")
        print(f"     └─ {err}")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()