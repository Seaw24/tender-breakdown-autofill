from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# CHANGE THIS to the actual path of your Master Breakdown file
file_path = "Tender Breakdown.xlsx"


def inspect_master_breakdown():
    try:
        # Load the workbook
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active  # Usually the first sheet

        print(f"\n{'='*80}")
        print(f"📊 INSPECTING BREAKDOWN MASTER: {file_path}")
        print(f"{'='*80}\n")

        # 1. IDENTIFY LOCATIONS (Row 1)
        print("📍 LOCATION COLUMN MAPPING (Row 1):")
        print(f"{'Location Name':<30} | {'Col Index':<10} | {'Letter'}")
        print("-" * 55)

        # We scan the first 200 columns
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().lower() != "location":
                letter = get_column_letter(cell.column)
                print(f"{str(cell.value):<30} | {cell.column:<10} | {letter}")

        # 2. IDENTIFY TENDER OFFSETS (Row 2)
        print("\n💰 TENDER COLUMN OFFSETS (Row 2):")
        print("Checking the first location block to find offsets...")
        print(f"{'Tender Name':<20} | {'Offset from Start'}")
        print("-" * 40)

        # We look at the first 12 columns to see the tender order
        for c in range(2, 14):
            val = ws.cell(2, c).value
            if val:
                print(f"{str(val):<20} | +{c - 2}")

        # 3. IDENTIFY DATE ROWS (Column A)
        print("\n📅 DATE ROW SAMPLES (Column A):")
        print(f"{'Row Number':<12} | {'Value'}")
        print("-" * 25)

        # Show a few rows to see where dates actually start
        for r in range(1, 10):
            val = ws.cell(r, 1).value
            print(f"Row {r:<8} | {val}")

        print(f"\n{'='*80}")

    except Exception as e:
        print(f"❌ Error: {e}")


if __name__ == "__main__":
    inspect_master_breakdown()
